import copy
import json
import re
import shutil
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from pydantic import BaseModel


DEFAULT_SHEET_MAPPING = {
    "sheet_name": "Sheet1",
    "report_date_cell": "C2",
    "data_table": {
        "header_row": 4,
        "data_start_row": 5,
        "data_end_row": 8,
        "columns": {
            "media": "A",
            "impression": "B",
            "click": "C",
            "cost": "D",
            "conversion": "E",
        },
    },
    "source_columns": {
        "media": "media",
        "impression": "impression",
        "click": "click",
        "cost": "cost",
        "conversion": "conversion",
    },
    "formula_columns": ["F", "G"],
    "total_row": 9,
    "note_cell": "A12",
}

DEFAULT_MAPPING = {"sheets": [DEFAULT_SHEET_MAPPING]}


class MappingColumns(BaseModel):
    media: str
    impression: str
    click: str
    cost: str
    conversion: str


class MappingTable(BaseModel):
    header_row: int
    data_start_row: int
    data_end_row: int
    columns: MappingColumns


class TemplateMapping(BaseModel):
    sheet_name: str
    report_date_cell: str
    data_table: MappingTable
    source_columns: MappingColumns
    formula_columns: list[str]
    total_row: int
    note_cell: str


class WorkbookMapping(BaseModel):
    sheets: list[TemplateMapping]


def extract_template_structure(
    workbook_bytes: bytes,
    max_rows: int | None = None,
    max_cols: int | None = None,
) -> dict[str, Any]:
    """Extract visible template structure for AI-assisted mapping."""
    wb = load_workbook(BytesIO(workbook_bytes), data_only=False)
    result: dict[str, Any] = {"sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info: dict[str, Any] = {
            "name": sheet_name,
            "dimensions": ws.dimensions,
            "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
            "cells": [],
        }

        scan_max_row = max_rows or ws.max_row
        scan_max_col = max_cols or ws.max_column

        for row in ws.iter_rows(min_row=1, max_row=scan_max_row, min_col=1, max_col=scan_max_col):
            for cell in row:
                if cell.value is None:
                    continue

                cell_info: dict[str, Any] = {
                    "addr": cell.coordinate,
                    "row": cell.row,
                    "col": cell.column,
                    "value": str(cell.value),
                }
                if cell.font and cell.font.bold:
                    cell_info["bold"] = True
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    rgb = cell.fill.start_color.rgb
                    if isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                        cell_info["fill"] = rgb
                if cell.number_format and cell.number_format != "General":
                    cell_info["fmt"] = cell.number_format
                sheet_info["cells"].append(cell_info)

        result["sheets"].append(sheet_info)

    return result


def build_ai_prompt(structure: dict[str, Any], data_preview: dict[str, Any] | None = None) -> str:
    """Build a prompt that asks an LLM to infer replaceable data locations."""
    sheet_blocks = []

    for index, sheet in enumerate(structure["sheets"], start=1):
        grid_lines = []
        for cell in sheet["cells"]:
            flags = []
            if cell.get("bold"):
                flags.append("bold")
            if cell.get("fill"):
                flags.append(f"fill={cell['fill'][-6:]}")
            if cell.get("fmt"):
                flags.append(f"fmt={cell['fmt']}")
            flag_str = f" [{', '.join(flags)}]" if flags else ""
            grid_lines.append(f"  {cell['addr']}: \"{cell['value']}\"{flag_str}")

        merged = "\n".join(f"  - {m}" for m in sheet["merged_ranges"]) or "  (없음)"
        cells = "\n".join(grid_lines) or "  (값이 있는 셀 없음)"
        sheet_blocks.append(
            f"""[시트 {index}]
시트명: {sheet['name']}
데이터 영역: {sheet['dimensions']}

병합된 셀:
{merged}

셀 값 및 서식:
{cells}"""
        )

    sheet_names = ", ".join(sheet["name"] for sheet in structure["sheets"])
    data_block = ""
    if data_preview:
        data_block = f"""

[업로드 데이터 컬럼]
{json.dumps(data_preview["columns"], ensure_ascii=False, indent=2)}

[업로드 데이터 샘플]
{json.dumps(data_preview["rows"], ensure_ascii=False, indent=2)}
"""

    return f"""다음은 광고 보고서 엑셀 템플릿의 전체 워크북 구조입니다.

[워크북 전체 시트 수] {len(structure["sheets"])}
[전체 시트] {sheet_names}

{chr(10).join(sheet_blocks)}
{data_block}

위 템플릿에서 매일 교체되어야 할 동적 데이터의 셀 위치를 시트별로 추론하여
업로드 데이터 컬럼과 템플릿 입력 필드도 함께 매핑해주세요.
sheets 배열에는 위에 나열된 모든 워크시트를 하나씩 포함하세요.
즉, 워크북 전체 시트 수가 4개이면 sheets 배열도 반드시 4개 항목이어야 합니다.
어떤 시트의 위치가 애매해도 가장 그럴듯한 데이터 영역을 추론해서 포함하세요.
아래 JSON 스키마로 답해주세요. 설명 없이 JSON만 출력하세요.

{{
  "sheets": [
    {{
      "sheet_name": "데이터를 교체할 시트명",
      "report_date_cell": "보고일자가 들어갈 셀 주소",
      "data_table": {{
        "header_row": 헤더 행 번호,
        "data_start_row": 데이터 시작 행,
        "data_end_row": 데이터 끝 행,
        "columns": {{
          "media": "매체명 컬럼 letter",
          "impression": "노출수 컬럼 letter",
          "click": "클릭수 컬럼 letter",
          "cost": "비용 컬럼 letter",
          "conversion": "전환수 컬럼 letter"
        }}
      }},
      "source_columns": {{
        "media": "업로드 데이터에서 매체명에 해당하는 컬럼명",
        "impression": "업로드 데이터에서 노출수에 해당하는 컬럼명",
        "click": "업로드 데이터에서 클릭수에 해당하는 컬럼명",
        "cost": "업로드 데이터에서 비용에 해당하는 컬럼명",
        "conversion": "업로드 데이터에서 전환수에 해당하는 컬럼명"
      }},
      "formula_columns": ["수식이 들어있어 건드리면 안 되는 컬럼 letter들"],
      "total_row": 합계 행 번호,
      "note_cell": "비고/메모 입력 셀 주소"
    }}
  ]
}}
"""


def generate_mapping_with_gemini(
    structure: dict[str, Any],
    api_key: str,
    model: str = "gemini-2.5-flash",
    data_preview: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Ask the Gemini API to infer a template mapping from extracted workbook structure."""
    try:
        from google import genai
    except ImportError as exc:
        raise ImportError(
            "Gemini SDK를 불러오지 못했습니다. Streamlit Cloud가 requirements.txt의 "
            "google-genai 패키지를 다시 설치하도록 앱을 재부팅하거나 redeploy하세요."
        ) from exc

    client = genai.Client(api_key=api_key)
    prompt = build_ai_prompt(structure, data_preview=data_preview)
    system_instruction = (
        "You infer Excel template mappings. Return only fields that identify where daily data "
        "should be written. Preserve formulas and total rows by excluding them from editable "
        "data columns. Return exactly one mapping item for every worksheet in the workbook. "
        "Map uploaded source data columns to the canonical fields: media, impression, click, "
        "cost, and conversion."
    )

    response = client.models.generate_content(
        model=model,
        contents=prompt,
        config={
            "system_instruction": system_instruction,
            "response_mime_type": "application/json",
            "response_schema": WorkbookMapping,
        },
    )

    parsed = response.parsed
    if isinstance(parsed, WorkbookMapping):
        return parsed.model_dump()
    return WorkbookMapping.model_validate_json(response.text).model_dump()


def read_data_table(uploaded_file: Any) -> pd.DataFrame:
    """Read daily replacement rows from CSV or Excel."""
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    return pd.read_excel(uploaded_file)


def normalize_daily_data(df: pd.DataFrame, report_date: str, note: str) -> dict[str, Any]:
    rows = df.where(pd.notnull(df), None).to_dict(orient="records")
    return {"report_date": report_date, "rows": rows, "note": note}


def build_data_preview(df: pd.DataFrame, max_rows: int = 5) -> dict[str, Any]:
    preview_df = df.head(max_rows).where(pd.notnull(df.head(max_rows)), None)
    return {
        "columns": [str(col) for col in df.columns],
        "rows": [
            {str(key): make_json_safe(value) for key, value in row.items()}
            for row in preview_df.to_dict(orient="records")
        ],
    }


def make_json_safe(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if hasattr(value, "item"):
        try:
            return value.item()
        except ValueError:
            pass
    return value


def lookup_row_value(row_data: dict[str, Any], source_column: str) -> Any:
    if source_column in row_data:
        return row_data[source_column]

    normalized_source = "".join(str(source_column).lower().split())
    for key, value in row_data.items():
        if "".join(str(key).lower().split()) == normalized_source:
            return value
    return None


def count_injectable_values(mapping: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    sheet_mappings = get_sheet_mappings(mapping)
    rows = data["rows"]
    count = 0
    missing_sources = set()
    sheet_counts = {}

    for sheet_mapping in sheet_mappings:
        sheet_count = 0
        cols = sheet_mapping["data_table"]["columns"]
        source_cols = sheet_mapping.get("source_columns", {})
        for row_data in rows:
            for field in cols:
                source_column = source_cols.get(field, field)
                value = lookup_row_value(row_data, source_column)
                if value is None:
                    missing_sources.add(source_column)
                else:
                    count += 1
                    sheet_count += 1
        sheet_counts[sheet_mapping.get("sheet_name", "(active sheet)")] = sheet_count

    return {
        "injectable_values": count,
        "missing_sources": sorted(str(source) for source in missing_sources),
        "sheet_counts": sheet_counts,
    }


def snapshot_design(workbook_bytes: bytes, max_rows: int = 40, max_cols: int = 15) -> dict[str, Any]:
    wb = load_workbook(BytesIO(workbook_bytes), data_only=False)
    snap: dict[str, Any] = {
        "sheet_names": wb.sheetnames,
        "sheets": {},
    }

    for ws in wb.worksheets:
        sheet_snap: dict[str, Any] = {
            "merged_ranges": sorted(str(r) for r in ws.merged_cells.ranges),
            "column_widths": {k: v.width for k, v in ws.column_dimensions.items() if v.width},
            "row_heights": {k: v.height for k, v in ws.row_dimensions.items() if v.height},
            "chart_count": len(ws._charts),
            "cells": {},
        }
        for row in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols):
            for cell in row:
                info: dict[str, Any] = {}
                if cell.font:
                    info["font_bold"] = cell.font.bold
                    info["font_color"] = (
                        str(cell.font.color.rgb) if cell.font.color and cell.font.color.type == "rgb" else None
                    )
                    info["font_name"] = cell.font.name
                    info["font_size"] = cell.font.size
                if cell.fill and cell.fill.start_color:
                    info["fill"] = str(cell.fill.start_color.rgb)
                if cell.border:
                    info["has_border"] = bool(cell.border.left and cell.border.left.style)
                if cell.number_format:
                    info["fmt"] = cell.number_format
                if cell.alignment:
                    info["align"] = (cell.alignment.horizontal, cell.alignment.vertical)
                sheet_snap["cells"][cell.coordinate] = info
        snap["sheets"][ws.title] = sheet_snap
    return snap


def compare_snapshots(before: dict[str, Any], after: dict[str, Any]) -> dict[str, Any]:
    structural_diffs = []
    if before["sheet_names"] != after["sheet_names"]:
        structural_diffs.append("시트 목록 또는 순서가 변경되었습니다.")

    style_changed = []
    for sheet_name, before_sheet in before["sheets"].items():
        after_sheet = after["sheets"].get(sheet_name)
        if not after_sheet:
            structural_diffs.append(f"{sheet_name} 시트가 사라졌습니다.")
            continue

        if before_sheet["merged_ranges"] != after_sheet["merged_ranges"]:
            structural_diffs.append(f"{sheet_name} 시트의 병합 영역이 변경되었습니다.")
        if before_sheet["column_widths"] != after_sheet["column_widths"]:
            structural_diffs.append(f"{sheet_name} 시트의 컬럼 너비가 변경되었습니다.")
        if before_sheet["row_heights"] != after_sheet["row_heights"]:
            structural_diffs.append(f"{sheet_name} 시트의 행 높이가 변경되었습니다.")
        if before_sheet["chart_count"] != after_sheet["chart_count"]:
            structural_diffs.append(f"{sheet_name} 시트의 차트 개수가 변경되었습니다.")

        for addr, before_info in before_sheet["cells"].items():
            after_info = after_sheet["cells"].get(addr, {})
            for key in ["font_bold", "font_color", "fill", "fmt", "align", "has_border"]:
                if before_info.get(key) != after_info.get(key):
                    style_changed.append(
                        {
                            "sheet": sheet_name,
                            "cell": addr,
                            "property": key,
                            "before": before_info.get(key),
                            "after": after_info.get(key),
                        }
                    )

    return {
        "structural_diffs": structural_diffs,
        "style_diffs_count": len(style_changed),
        "style_diffs_sample": style_changed[:10],
    }


def inject_data(template_bytes: bytes, mapping: dict[str, Any], data: dict[str, Any]) -> bytes:
    """Copy the template and replace values only, preserving styles, formulas, and charts."""
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as template_file:
        template_path = Path(template_file.name)
        template_file.write(template_bytes)

    output_path = template_path.with_name(f"{template_path.stem}_output.xlsx")
    try:
        shutil.copy(template_path, output_path)
        wb = load_workbook(output_path, data_only=False)

        for sheet_mapping in get_sheet_mappings(mapping):
            inject_sheet_data(wb, sheet_mapping, data)

        wb.save(output_path)
        return output_path.read_bytes()
    finally:
        template_path.unlink(missing_ok=True)
        output_path.unlink(missing_ok=True)


def inject_sheet_data(wb: Any, sheet_mapping: dict[str, Any], data: dict[str, Any]) -> None:
    sheet_name = sheet_mapping.get("sheet_name")
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    if sheet_mapping.get("report_date_cell"):
        set_cell_value(ws, sheet_mapping["report_date_cell"], data.get("report_date"))

    table = sheet_mapping["data_table"]
    cols = table["columns"]
    source_cols = sheet_mapping.get("source_columns", {})
    start_row = int(table["data_start_row"])
    end_row = int(table["data_end_row"])

    for index, row_data in enumerate(data["rows"]):
        excel_row = start_row + index
        if excel_row > end_row:
            copy_row_template(ws, end_row, excel_row)
        for field, col_letter in cols.items():
            source_column = source_cols.get(field, field)
            value = lookup_row_value(row_data, source_column)
            if value is not None:
                set_cell_value(ws, f"{col_letter}{excel_row}", value)

    if sheet_mapping.get("note_cell"):
        set_cell_value(ws, sheet_mapping["note_cell"], data.get("note"))


def set_cell_value(ws: Any, cell_addr: str, value: Any) -> None:
    cell = ws[cell_addr]
    if isinstance(cell, MergedCell):
        cell = ws[get_merged_parent_address(ws, cell_addr)]
    cell.value = value


def copy_row_template(ws: Any, source_row: int, target_row: int) -> None:
    if target_row <= source_row:
        return

    source_dimension = ws.row_dimensions[source_row]
    target_dimension = ws.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden

    for col_index in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col_index)
        target_cell = ws.cell(row=target_row, column=col_index)

        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy.copy(source_cell.protection)

        if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
            target_cell.value = translate_formula(source_cell.value, source_cell.coordinate, target_cell.coordinate)


def translate_formula(formula: str, origin: str, target: str) -> str:
    try:
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        return formula


def get_merged_parent_address(ws: Any, cell_addr: str) -> str:
    for merged_range in ws.merged_cells.ranges:
        if cell_addr in merged_range:
            return merged_range.start_cell.coordinate
    return cell_addr


def validate_mapping(mapping_text: str) -> dict[str, Any]:
    cleaned_text = clean_mapping_text(mapping_text)
    if not cleaned_text:
        raise ValueError("매핑 JSON이 비어 있습니다. Gemini API로 매핑을 생성하거나 JSON을 직접 입력하세요.")

    try:
        mapping = json.loads(cleaned_text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"매핑 JSON 형식이 올바르지 않습니다: {exc.msg}") from exc

    sheet_mappings = get_sheet_mappings(mapping)
    if not sheet_mappings:
        raise ValueError("매핑 JSON의 sheets 배열이 비어 있습니다.")

    for index, sheet_mapping in enumerate(sheet_mappings, start=1):
        required = ["sheet_name", "report_date_cell", "data_table", "note_cell"]
        missing = [key for key in required if key not in sheet_mapping]
        if missing:
            raise ValueError(f"{index}번째 시트 매핑에 필수 키가 없습니다: {', '.join(missing)}")
        if "columns" not in sheet_mapping["data_table"]:
            raise ValueError(f"{index}번째 시트 매핑에 data_table.columns가 필요합니다.")
        if "source_columns" not in sheet_mapping:
            sheet_mapping["source_columns"] = {
                field: field for field in sheet_mapping["data_table"]["columns"]
            }

    if "sheets" not in mapping:
        mapping = {"sheets": sheet_mappings}
    return mapping


def get_sheet_mappings(mapping: dict[str, Any]) -> list[dict[str, Any]]:
    if "sheets" in mapping:
        return mapping.get("sheets") or []
    return [mapping]


def clean_mapping_text(mapping_text: str | None) -> str:
    text = (mapping_text or "").strip()
    if not text:
        return ""

    fenced = re.search(r"```(?:json)?\s*(.*?)```", text, flags=re.IGNORECASE | re.DOTALL)
    if fenced:
        text = fenced.group(1).strip()

    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        text = text[start : end + 1]

    return text.strip()


def mapping_as_text(mapping: dict[str, Any] | None = None) -> str:
    return json.dumps(copy.deepcopy(mapping or DEFAULT_MAPPING), ensure_ascii=False, indent=2)
